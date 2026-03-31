#!/usr/bin/env python3
"""
SWRCB eAR Bulk Rate Ingest

Purpose:
    Loads water rate data from the California State Water Resources Control
    Board's Electronic Annual Report (eAR), cleaned and formatted by
    HydroShare (Erik Porse, UC ANR / California Institute for Water Resources).

    Maps eAR single-family residential rate tier structure to the water_rates
    schema. Only ingests PWSIDs that already exist in our cws_boundaries table
    (i.e., systems we're tracking via MDWD or other sources).

    Supports multiple years (2020-2022). Each year is stored as a separate
    record with source='swrcb_ear_YYYY'. Existing records for the same
    PWSID + year + source are replaced (idempotent).

Author: AI-Generated
Created: 2026-03-24
Modified: 2026-03-24

Dependencies:
    - openpyxl (Excel reading)
    - sqlalchemy (database)
    - loguru (logging)

Usage:
    # CLI
    ua-ingest ear                          # Ingest all available years
    ua-ingest ear --year 2022              # Single year
    ua-ingest ear --year 2022 --dry-run    # Preview without DB writes

    # Python
    from utility_api.ingest.ear_ingest import run_ear_ingest
    run_ear_ingest(years=[2022])

Notes:
    - HCF (Hundred Cubic Feet) = CCF. No unit conversion needed.
    - eAR bill columns (WR6/9/12/24HCFDWCharges) are monthly-equivalent.
    - For bimonthly billing: base + volumetric is per-billing-period; the
      WR*HCFDWCharges columns are already divided by 2 for monthly.
    - Fixed charge normalization: CostPerUOM1 is per-billing-period.
      We divide by 2 for bimonthly, by 3 for quarterly.
    - Rate structure mapping:
        eAR 'Variable Base'   -> 'increasing_block' (tiered volumetric)
        eAR 'Uniform Usage'   -> 'uniform' (single volumetric rate)
        eAR 'Fixed Base'      -> 'flat' (no volumetric component)
        eAR 'OtherRate'       -> 'other' (budget-based, seasonal, etc.)

Data Sources:
    - Input: data/raw/swrcb_ear/ear_annual_matrix_YYYY.xlsx (HydroShare)
    - Input: utility.cws_boundaries (PWSID filter)
    - Output: utility.water_rates table (source=swrcb_ear_YYYY)

Configuration:
    - Excel files must be in data/raw/swrcb_ear/
    - Database connection via .env (DATABASE_URL)
"""

from datetime import date, datetime, timezone
from pathlib import Path

from loguru import logger
from sqlalchemy import text

from utility_api.config import settings
from utility_api.db import engine
from utility_api.ops.rate_schedule_helpers import water_rate_to_schedule, write_rate_schedule


# --- Column name mapping ---
# Maps our internal keys to eAR Excel header names.
# Column positions vary between years (2020: 1314 cols, 2021: 2315, 2022: 2978)
# so we look up by header name at load time, not by hardcoded index.

COL_NAMES = {
    "pwsid": "PwsID",
    "pws_name": "PWSName",
    "population": "PopulationTotal",
    "billing_freq": "WRBillingFreq",
    "rate_structure": "WRRateStructureRes",
    "uom": "WRUOM",
    "has_rate": "WRHasRate",
    "sf_num_tiers": "WRSFNumTiers",
    # Single-family tier structure (up to 4 tiers)
    "sf_base_1": "WRSFCostPerUOM1",       # Fixed/base charge per billing period
    "sf_limit_1": "WRSFMetricUsage1",     # Tier 1 upper limit (HCF/billing period)
    "sf_rate_1": "WRSFUsageCost1",        # $/HCF for tier 1
    "sf_base_2": "WRSFCostPerUOM2",
    "sf_limit_2": "WRSFMetricUsage2",
    "sf_rate_2": "WRSFUsageCost2",
    "sf_base_3": "WRSFCostPerUOM3",
    "sf_limit_3": "WRSFMetricUsage3",
    "sf_rate_3": "WRSFUsageCost3",
    "sf_base_4": "WRSFCostPerUOM4",
    "sf_limit_4": "WRSFMetricUsage4",
    "sf_rate_4": "WRSFUsageCost4",
    # Provenance
    "rate_updated": "WRRateUpdatedDate",
    "rate_link": "WRWaterRateLink",
    # Pre-computed monthly-equivalent bills (not present in 2020)
    "bill_6hcf": "WR6HCFDWCharges",
    "bill_9hcf": "WR9HCFDWCharges",
    "bill_12hcf": "WR12HCFDWCharges",
    "bill_24hcf": "WR24HCFDWCharges",
}


def _build_col_index(headers: list) -> dict[str, int | None]:
    """Build column index mapping from Excel header row.

    Looks up each COL_NAMES entry by header name and returns its 0-based index.
    Returns None for columns not found in this year's file.
    """
    header_map = {}
    for i, h in enumerate(headers):
        if h is not None:
            header_map[h] = i

    col = {}
    for key, header_name in COL_NAMES.items():
        col[key] = header_map.get(header_name)

    # Log any missing columns
    missing = [k for k, v in col.items() if v is None]
    if missing:
        logger.warning(f"  Missing columns in this year's file: {missing}")

    return col

# eAR rate structure -> our normalized type
STRUCTURE_MAP = {
    "Variable Base": "increasing_block",
    "Uniform Usage": "uniform",
    "Fixed Base": "flat",
    "OtherRate": "other",
}

# Billing frequency normalization
BILLING_FREQ_MAP = {
    "M": "monthly",
    "BM": "bimonthly",
    "Q": "quarterly",
    "A": "annually",
}

# Divisor for normalizing per-billing-period charges to monthly
BILLING_DIVISOR = {
    "M": 1,
    "BM": 2,
    "Q": 3,
    "A": 12,
}

# Data directory
DATA_DIR = Path(__file__).resolve().parents[3] / "data" / "raw" / "swrcb_ear"

# Available years and their filenames
YEAR_FILES = {
    2020: "ear_annual_matrix_2020.xlsx",
    2021: "ear_annual_matrix_2021.xlsx",
    2022: "ear_annual_matrix_2022.xlsx",
}


def _safe_float(val) -> float | None:
    """Convert a cell value to float, returning None for empty/invalid."""
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _safe_int(val) -> int | None:
    """Convert a cell value to int, returning None for empty/invalid."""
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _extract_date(val) -> date | None:
    """Extract a date from a datetime cell value."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    # Try parsing string
    try:
        return datetime.strptime(str(val).strip(), "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _get_existing_pwsids() -> set[str]:
    """Get the set of CA PWSIDs that exist in our database (cws_boundaries joined to mdwd)."""
    schema = settings.utility_schema
    with engine.connect() as conn:
        rows = conn.execute(text(f"""
            SELECT c.pwsid
            FROM {schema}.cws_boundaries c
            INNER JOIN {schema}.mdwd_financials m ON m.pwsid = c.pwsid
            WHERE c.state_code = 'CA'
        """)).fetchall()
    return {r[0] for r in rows}


def _get_cell(row: tuple, col: dict, key: str):
    """Safely get a cell value using the dynamic column index.

    Returns None if the column doesn't exist in this year's file.
    """
    idx = col.get(key)
    if idx is None:
        return None
    if idx >= len(row):
        return None
    return row[idx]


def _parse_ear_row(row: tuple, year: int, col: dict) -> dict | None:
    """Parse a single eAR row into a water_rates-compatible dict.

    Parameters
    ----------
    row : tuple
        Raw row values from the Excel sheet.
    year : int
        Report year (2020, 2021, or 2022).
    col : dict
        Column index mapping (from _build_col_index).

    Returns
    -------
    dict | None
        Mapped record ready for DB insert, or None if row should be skipped.
    """
    pwsid = _get_cell(row, col, "pwsid")
    has_rate = _get_cell(row, col, "has_rate")

    # Skip rows without rate data
    if not pwsid or has_rate != "Yes":
        return None

    # Rate structure
    raw_structure = _get_cell(row, col, "rate_structure")
    structure_type = STRUCTURE_MAP.get(raw_structure, "other") if raw_structure else None

    # Billing frequency
    raw_freq = _get_cell(row, col, "billing_freq")
    billing_freq = BILLING_FREQ_MAP.get(raw_freq)
    divisor = BILLING_DIVISOR.get(raw_freq, 1)

    # Fixed charge — CostPerUOM1 is per-billing-period, normalize to monthly
    raw_base = _safe_float(_get_cell(row, col, "sf_base_1"))
    fixed_charge_monthly = round(raw_base / divisor, 2) if raw_base is not None else None

    # Tier limits and rates (HCF per billing period → HCF per month)
    # Tier limits are per-billing-period; normalize to monthly for consistency
    def tier_limit(col_key: str) -> float | None:
        val = _safe_float(_get_cell(row, col, col_key))
        if val is not None:
            return round(val / divisor, 2)
        return None

    def tier_rate(col_key: str) -> float | None:
        return _safe_float(_get_cell(row, col, col_key))

    # Effective date: use WRRateUpdatedDate if available, else Jan 1 of report year
    eff_date = _extract_date(_get_cell(row, col, "rate_updated"))
    if eff_date is None:
        eff_date = date(year, 1, 1)

    # Bill snapshots (already monthly-equivalent in eAR; not present in 2020)
    bill_6 = _safe_float(_get_cell(row, col, "bill_6hcf"))
    bill_9 = _safe_float(_get_cell(row, col, "bill_9hcf"))
    bill_12 = _safe_float(_get_cell(row, col, "bill_12hcf"))
    bill_24 = _safe_float(_get_cell(row, col, "bill_24hcf"))

    # Determine confidence based on data completeness
    has_bills = any(v is not None for v in [bill_6, bill_9, bill_12, bill_24])
    has_tiers = _safe_float(_get_cell(row, col, "sf_rate_1")) is not None
    if has_bills and has_tiers:
        confidence = "high"
    elif has_bills:
        confidence = "medium"
    elif has_tiers:
        confidence = "medium"
    else:
        confidence = "low"

    # Build notes
    notes_parts = []
    if raw_structure and raw_structure not in STRUCTURE_MAP:
        notes_parts.append(f"Unmapped rate structure: {raw_structure}")
    num_tiers = _safe_int(_get_cell(row, col, "sf_num_tiers"))
    if num_tiers and num_tiers > 4:
        notes_parts.append(f"eAR reports {num_tiers} tiers; only first 4 captured")
    uom = _get_cell(row, col, "uom")
    if uom and "Hundred Cubic Feet" not in str(uom):
        notes_parts.append(f"Non-standard UOM: {uom}")

    return {
        "pwsid": pwsid,
        "utility_name": _get_cell(row, col, "pws_name"),
        "state_code": "CA",
        "county": None,  # eAR doesn't include county in a clean column
        "rate_effective_date": eff_date,
        "rate_structure_type": structure_type,
        "rate_class": "residential",
        "billing_frequency": billing_freq,
        "fixed_charge_monthly": fixed_charge_monthly,
        "meter_size_inches": None,  # eAR doesn't report meter size
        "tier_1_limit_ccf": tier_limit("sf_limit_1"),
        "tier_1_rate": tier_rate("sf_rate_1"),
        "tier_2_limit_ccf": tier_limit("sf_limit_2"),
        "tier_2_rate": tier_rate("sf_rate_2"),
        "tier_3_limit_ccf": tier_limit("sf_limit_3"),
        "tier_3_rate": tier_rate("sf_rate_3"),
        "tier_4_limit_ccf": tier_limit("sf_limit_4"),
        "tier_4_rate": tier_rate("sf_rate_4"),
        "bill_5ccf": None,   # eAR doesn't report at 5 CCF
        "bill_10ccf": None,  # eAR doesn't report at 10 CCF
        "bill_6ccf": bill_6,
        "bill_9ccf": bill_9,
        "bill_12ccf": bill_12,
        "bill_24ccf": bill_24,
        "source": f"swrcb_ear_{year}",
        "source_url": _get_cell(row, col, "rate_link") or None,
        "raw_text_hash": None,
        "parse_confidence": confidence,
        "parse_model": None,
        "parse_notes": "; ".join(notes_parts) if notes_parts else None,
    }


def _load_ear_year(year: int, target_pwsids: set[str], dry_run: bool = False) -> dict:
    """Load a single year of eAR data.

    Parameters
    ----------
    year : int
        Report year (2020, 2021, or 2022).
    target_pwsids : set[str]
        PWSIDs to include (intersection with our DB).
    dry_run : bool
        If True, parse but don't write to DB.

    Returns
    -------
    dict
        Stats: total_rows, matched, with_rates, inserted, skipped.
    """
    import openpyxl

    filename = YEAR_FILES.get(year)
    if filename is None:
        logger.error(f"No file mapping for year {year}")
        return {"error": f"Unknown year: {year}"}

    filepath = DATA_DIR / filename
    if not filepath.exists():
        logger.error(f"File not found: {filepath}")
        logger.info(f"Download from: https://www.hydroshare.org/resource/8108599db4934252a5d0e6e83b5d3551/")
        return {"error": f"File not found: {filepath}"}

    logger.info(f"Loading eAR {year}: {filepath.name} ({filepath.stat().st_size / 1024 / 1024:.1f} MB)")

    wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
    ws = wb["ear_annual_matrix"]

    # Build column index from header row
    headers = None
    for row in ws.iter_rows(max_row=1, values_only=True):
        headers = list(row)
        break

    if headers is None:
        logger.error(f"  No header row found in {filepath.name}")
        wb.close()
        return {"error": "No header row"}

    col = _build_col_index(headers)

    # Verify essential columns exist
    if col.get("pwsid") is None:
        logger.error(f"  PwsID column not found in {filepath.name}")
        wb.close()
        return {"error": "PwsID column missing"}

    logger.info(f"  Column mapping built: {len(headers)} columns, PwsID at index {col['pwsid']}")

    stats = {
        "year": year,
        "total_rows": 0,
        "matched": 0,
        "with_rates": 0,
        "inserted": 0,
        "skipped_no_rate": 0,
        "skipped_low_confidence": 0,
    }

    records = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        stats["total_rows"] += 1
        pwsid = _get_cell(row, col, "pwsid")

        if pwsid not in target_pwsids:
            continue
        stats["matched"] += 1

        record = _parse_ear_row(row, year, col)
        if record is None:
            stats["skipped_no_rate"] += 1
            continue
        stats["with_rates"] += 1

        records.append(record)

    wb.close()

    logger.info(
        f"  eAR {year}: {stats['total_rows']} total systems, "
        f"{stats['matched']} matched our DB, "
        f"{stats['with_rates']} have rate data"
    )

    if dry_run:
        logger.info(f"  [DRY RUN] Would insert {len(records)} records")
        # Print a few samples
        for r in records[:5]:
            logger.info(
                f"    {r['pwsid']} | {r['utility_name'][:40]} | "
                f"{r['rate_structure_type']} | "
                f"fixed=${r['fixed_charge_monthly'] or 0:.2f} | "
                f"bill@6=${r['bill_6ccf'] or 0:.2f} | "
                f"bill@12=${r['bill_12ccf'] or 0:.2f} | "
                f"[{r['parse_confidence']}]"
            )
        if len(records) > 5:
            logger.info(f"    ... and {len(records) - 5} more")
        stats["inserted"] = 0
        return stats

    # Write to database
    schema = settings.utility_schema
    source_tag = f"swrcb_ear_{year}"

    with engine.connect() as conn:
        # Delete existing eAR records for this year (idempotent re-run)
        deleted = conn.execute(text(f"""
            DELETE FROM {schema}.rate_schedules
            WHERE source_key = :source
        """), {"source": source_tag}).rowcount
        if deleted:
            logger.info(f"  Cleared {deleted} existing {source_tag} records from rate_schedules")

        # Batch insert into rate_schedules (Phase 3: direct write, no water_rates)
        for record in records:
            schedule = water_rate_to_schedule(record)
            write_rate_schedule(conn, schedule)

        conn.commit()
        stats["inserted"] = len(records)

    logger.info(f"  Inserted {stats['inserted']} records for eAR {year}")
    return stats


def run_ear_ingest(
    years: list[int] | None = None,
    dry_run: bool = False,
) -> dict:
    """Run the SWRCB eAR bulk rate ingest.

    Parameters
    ----------
    years : list[int] | None
        Years to ingest. If None, ingests all available (2020-2022).
    dry_run : bool
        If True, parse and report but don't write to DB.

    Returns
    -------
    dict
        Summary stats across all years.
    """
    started = datetime.now(timezone.utc)
    logger.info("=== SWRCB eAR Rate Ingest Starting ===")

    if years is None:
        # Default: only ingest years whose files exist
        years = [y for y, f in YEAR_FILES.items() if (DATA_DIR / f).exists()]
        if not years:
            logger.error(f"No eAR Excel files found in {DATA_DIR}")
            logger.info("Download from: https://www.hydroshare.org/resource/8108599db4934252a5d0e6e83b5d3551/")
            return {"error": "No data files found"}

    logger.info(f"Years to ingest: {years}")

    # Get target PWSIDs from our database
    target_pwsids = _get_existing_pwsids()
    logger.info(f"Target CA PWSIDs in our DB: {len(target_pwsids)}")

    if not target_pwsids:
        logger.warning("No CA PWSIDs found in cws_boundaries + mdwd_financials. Nothing to ingest.")
        return {"error": "No target PWSIDs"}

    # Process each year
    all_stats = []
    for year in sorted(years):
        year_stats = _load_ear_year(year, target_pwsids, dry_run=dry_run)
        all_stats.append(year_stats)

    # Summary
    total_inserted = sum(s.get("inserted", 0) for s in all_stats)
    total_matched = sum(s.get("matched", 0) for s in all_stats)
    total_with_rates = sum(s.get("with_rates", 0) for s in all_stats)

    elapsed = (datetime.now(timezone.utc) - started).total_seconds()
    logger.info(f"\n=== eAR Ingest Complete ({elapsed:.0f}s) ===")
    logger.info(f"  Years processed: {years}")
    logger.info(f"  PWSIDs matched: {total_matched}")
    logger.info(f"  With rate data: {total_with_rates}")
    logger.info(f"  Records inserted: {total_inserted}")

    # Log pipeline run
    if not dry_run and total_inserted > 0:
        schema = settings.utility_schema
        with engine.connect() as conn:
            conn.execute(text(f"""
                INSERT INTO {schema}.pipeline_runs
                    (step_name, started_at, finished_at, row_count, status, notes)
                VALUES (:step, :started, NOW(), :count, 'success', :notes)
            """), {
                "step": "ear-ingest",
                "started": started,
                "count": total_inserted,
                "notes": (
                    f"years={years}, matched={total_matched}, "
                    f"with_rates={total_with_rates}, inserted={total_inserted}"
                ),
            })
            conn.commit()

    return {
        "years": years,
        "matched": total_matched,
        "with_rates": total_with_rates,
        "inserted": total_inserted,
        "per_year": all_stats,
    }
